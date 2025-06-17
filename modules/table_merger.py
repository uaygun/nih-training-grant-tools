import pandas as pd
from google.colab import files
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins

# Upload Excel files
uploaded = files.upload()

# Initialize containers for each table
combined_table2, combined_table4 = [], []
combined_table5a, combined_table5b = [], []
combined_table8a, combined_table8c = [], []

# Read all relevant sheets from each file
for filename in uploaded:
    excel_file = pd.ExcelFile(io.BytesIO(uploaded[filename]))

    if 'Table 2' in excel_file.sheet_names:
        df = excel_file.parse('Table 2').dropna(how='all')
        combined_table2.append(df)

    if 'Table 4' in excel_file.sheet_names:
        df = excel_file.parse('Table 4').dropna(how='all')
        combined_table4.append(df)

    if 'Table 5A' in excel_file.sheet_names:
        df = excel_file.parse('Table 5A').dropna(how='all')
        df = df.iloc[2:]  # Skip second row (row index 1)
        combined_table5a.append(df)

    if 'Table 5B' in excel_file.sheet_names:
        df = excel_file.parse('Table 5B').dropna(how='all')
        df = df.iloc[2:]  # Skip second row
        combined_table5b.append(df)

    if 'Table 8A' in excel_file.sheet_names:
        df = excel_file.parse('Table 8A').dropna(how='all')
        combined_table8a.append(df)

    if 'Table 8C' in excel_file.sheet_names:
        df = excel_file.parse('Table 8C').dropna(how='all')
        combined_table8c.append(df)

# Save merged output
output_filename = "Formatted_Merged_Tables.xlsx"
with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    if combined_table2:
        pd.concat(combined_table2, ignore_index=True).to_excel(writer, sheet_name="Table 2", index=False)
    if combined_table4:
        pd.concat(combined_table4, ignore_index=True).to_excel(writer, sheet_name="Table 4", index=False)
    if combined_table5a:
        pd.concat(combined_table5a, ignore_index=True).to_excel(writer, sheet_name="Table 5A", index=False)
    if combined_table5b:
        pd.concat(combined_table5b, ignore_index=True).to_excel(writer, sheet_name="Table 5B", index=False)
    if combined_table8a:
        pd.concat(combined_table8a, ignore_index=True).to_excel(writer, sheet_name="Table 8A", index=False)
    if combined_table8c:
        pd.concat(combined_table8c, ignore_index=True).to_excel(writer, sheet_name="Table 8C", index=False)

# Format function
def format_sheet(ws, column_widths, row1_height=70, other_row_height=50):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row[0].row].height = row1_height if row[0].row == 1 else other_row_height

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

# Load workbook
wb = load_workbook(output_filename)

# Format each table sheet
if "Table 2" in wb.sheetnames:
    format_sheet(wb["Table 2"], {
        'D': 20, 'E': 40,
        'G': 13, 'H': 13, 'I': 13, 'J': 13, 'K': 13, 'L': 13
    })

if "Table 4" in wb.sheetnames:
    ws = wb["Table 4"]
    format_sheet(ws, {
        'A': 20, 'B': 15, 'C': 20, 'D': 15, 'E': 40, 'F': 20, 'G': 20
    })
    # Currency formatting for column G
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7, max_row=ws.max_row):
        for cell in row:
            cell.number_format = '"$"#,##0'

if "Table 5A" in wb.sheetnames:
    format_sheet(wb["Table 5A"], {
        'A': 20, 'B': 25, 'C': 15, 'D': 20, 'E': 60
    })

if "Table 5B" in wb.sheetnames:
    format_sheet(wb["Table 5B"], {
        'A': 20, 'B': 25, 'C': 15, 'D': 20, 'E': 60
    })

if "Table 8A" in wb.sheetnames:
    format_sheet(wb["Table 8A"], {
        'A': 20, 'B': 20, 'C': 22, 'D': 13, 'E': 25, 'F': 25, 'G': 40,
        'H': 30, 'I': 30, 'J': 30
    })

if "Table 8C" in wb.sheetnames:
    format_sheet(wb["Table 8C"], {
        'A': 20, 'B': 20, 'C': 22, 'D': 13, 'E': 25, 'F': 25, 'G': 40,
        'H': 30, 'I': 30, 'J': 30
    })

# Save and download
wb.save(output_filename)
files.download(output_filename)
